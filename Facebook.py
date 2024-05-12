import os
import openpyxl
import datetime
from tkinter import *
from tkinter import messagebox
from functools import partial
Today = datetime.datetime.now()
Date = Today.strftime("%d-%m-%Y")

file_path = 'Project2.xlsx'

# Check if the file exists
if os.path.exists(file_path):
    Workbook1 = openpyxl.load_workbook(file_path)
else:
    # Create a new workbook
    Workbook1 = openpyxl.Workbook()

# Create sheets if they don't exist
for sheet_name in ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5', 'Sheet6',
                    'Sheet7', 'Sheet8', 'Sheet9', 'Sheet10', 'Sheet11', 'Sheet12', 'Sheet13', 'Sheet14']:
    if sheet_name not in Workbook1.sheetnames:
        Workbook1.create_sheet(title=sheet_name)

# Now you can access the sheets
WorkSheet1 = Workbook1['Sheet1']
WorkSheet2 = Workbook1['Sheet2']
WorkSheet3 = Workbook1['Sheet3']
WorkSheet4 = Workbook1['Sheet4']
WorkSheet5 = Workbook1['Sheet5']
WorkSheet6 = Workbook1['Sheet6']
WorkSheet7 = Workbook1['Sheet7']
WorkSheet8 = Workbook1['Sheet8']
WorkSheet9 = Workbook1['Sheet9']
WorkSheet10 = Workbook1['Sheet10']
WorkSheet11 = Workbook1['Sheet11']
WorkSheet12 = Workbook1['Sheet12']
WorkSheet13 = Workbook1['Sheet13']
WorkSheet14 = Workbook1['Sheet14']

# Save the workbook
Workbook1.save(file_path)



def Headersaving(worksheet, list):
    for heading in range(len(list)):
        worksheet.cell(row=1, column=heading + 1).value = list[heading]
        Workbook1.save("Project2.xlsx")

headersheet1 = ['Name', '', 'Username', 'Gmails', '', 'Joining Date', 'Gender', 'Education', 'Work', 'City', 'Country','', 'Bio']
headersheet2 = ['Name', 'Password', 'Gmail', 'Username', 'Comment', 'Friend_request', 'Login person']
headersheet3 = ['Temporary Friend-requests']
headersheet4 = ['Permanent Friends']
headersheet5 = ['Username', 'Post']
headersheet6 = ['Message', 'Sender', 'Receiver', '', '', 'Message', 'Sender', 'Receiver']
headersheet7 = ['Privacy_lists for posts']
headersheet8 = ['Block_Members lists']
Headersaving(WorkSheet1, headersheet1)
Headersaving(WorkSheet2, headersheet2)
Headersaving(WorkSheet3, headersheet3)
Headersaving(WorkSheet4, headersheet4)
Headersaving(WorkSheet5, headersheet5)
Headersaving(WorkSheet6, headersheet6)
Headersaving(WorkSheet7, headersheet7)
Headersaving(WorkSheet8, headersheet8)

class Helperfunctions:
    def Appendsheetall(worksheet, givendata, columnnumber):
        rownumber = 1
        while worksheet.cell(row=rownumber, column=columnnumber).value != None:
            rownumber += 1
        worksheet.cell(row=rownumber, column=columnnumber).value = givendata
        Workbook1.save("Project2.xlsx")
    def Userbio(bio, colomnnumber, Index):
        o = 0
        for r in WorkSheet1[colomnnumber]:
            o += 1
            if o == Index:
                print(bio, ":", r.value)
    def Commentstatus(a, Index):
        o = 0
        for r in WorkSheet2['E']:
            o += 1
            if o == Index:
                WorkSheet2.cell(row=Index, column=5).value = a
                Workbook1.save("Project2.xlsx")
    def Appendforfriends(worksheet, givendata, rownumber):
        columnnumber = 1
        while worksheet.cell(row=rownumber, column=columnnumber).value != None:
            columnnumber += 1
        worksheet.cell(row=rownumber, column=columnnumber).value = givendata
        Workbook1.save("Project2.xlsx")
    def Friendrequeststatus(a, Index):
        o = 0
        for r in WorkSheet2['F']:
            o += 1
            if o == Index:
                WorkSheet2.cell(row=Index, column=6).value = a
                Workbook1.save("Project2.xlsx")
    def Checking_blocking(check):
        WorkSheet2.cell(row=3, column=7).value = 1
        for a in range(len(WorkSheet8['1'])):
            if WorkSheet8['1'][a].value == check:
                k = 2
                while k <= WorkSheet8.max_row:
                    if WorkSheet8.cell(row=k, column=a + 1).value == WorkSheet2.cell(row=2, column=7).value:
                        WorkSheet2.cell(row=3, column=7).value = 2
                        break
                    k += 1
        Workbook1.save("Project2.xlsx")
    def You_Blocked(check):
        WorkSheet2.cell(row=4, column=7).value = 1
        for a in range(len(WorkSheet8['1'])):
            if WorkSheet8['1'][a].value == WorkSheet2.cell(row=2, column=7).value:
                k = 2
                while k <= WorkSheet8.max_row:
                    if WorkSheet8.cell(row=k, column=a + 1).value == check:
                        WorkSheet2.cell(row=4, column=7).value = 2
                        break
                    k += 1
        Workbook1.save("Project2.xlsx")
    def OFF_searching(self, off):
        off.destroy()
        Program_execution()

class Facebook:
    def Signup(self, Name, Password, Mail, Username, TK):
        check = False
        for i in range(len(WorkSheet2['D'])):
            if WorkSheet2.cell(row=i + 1, column=4).value == Username.get():
                check = True
                break
        if check == True:
            messagebox.showinfo("Warning", "This username is already exist.")
            messagebox.showinfo("Try Again", "Please choose another username !")
        else:
            Helperfunctions.Appendsheetall(WorkSheet2, Name.get(), 1)
            Helperfunctions.Appendsheetall(WorkSheet2, Password.get(), 2)
            Helperfunctions.Appendsheetall(WorkSheet2, Mail.get(), 3)
            Helperfunctions.Appendsheetall(WorkSheet2, Username.get(), 4)
            messagebox.showinfo("Sign_up",Username.get()+" sign-up Successfully.")
            TK.destroy()
            Program_execution()
    def Login(self, username, password, loginin):
        Check = 0
        Index_check = 0
        for i in range(len(WorkSheet2['D'])):
            Index_check += 1
            if username.get() == WorkSheet2['D'][i].value and password.get() == WorkSheet2['B'][i].value:
                messagebox.showinfo(username.get(),"Login successfully !")
                WorkSheet2.cell(row=2, column=7).value = username.get()
                Workbook1.save("Project2.xlsx")
                for n in range(len(WorkSheet2['D'])):
                    if username.get() == WorkSheet2['D'][n].value:
                        if WorkSheet2.cell(row=n + 1, column=5).value != None and WorkSheet2.cell(row=n + 1,column=5).value == True:
                            messagebox.showinfo("Notification", "You have new comment on your post.")
                        if WorkSheet2.cell(row=n + 1, column=6).value != None and WorkSheet2.cell(row=n + 1,column=6).value == True:
                            messagebox.showinfo("Notification", "You have new friend request.")
                for z in range(len(WorkSheet6['C'])):
                    if WorkSheet6['C'][z].value == username.get():
                        messagebox.showinfo("Notification","You have new Message from your friend.")
                        break
                Check += 1
                login = 0
                for k in range(len(WorkSheet1['C'])):
                    if username.get() == WorkSheet1['C'][k].value:
                        login += 1
                        break
                if login == 0:
                    Helperfunctions.Appendsheetall(WorkSheet1,Date,6)
                    Helperfunctions.Appendsheetall(WorkSheet1, username.get(), 3)
                    Helperfunctions.Appendforfriends(WorkSheet3, username.get(), 1)
                    Helperfunctions.Appendforfriends(WorkSheet4, username.get(), 1)
                    Helperfunctions.Appendsheetall(WorkSheet5, username.get(), 1)
                    Helperfunctions.Appendforfriends(WorkSheet7, username.get(), 1)
                    Helperfunctions.Appendforfriends(WorkSheet8, username.get(), 1)
                    p = 0
                    for m in WorkSheet2['A']:
                        p += 1
                        if p == Index_check:
                            Helperfunctions.Appendsheetall(WorkSheet1, m.value, 1)
                    o = 0
                    for r in WorkSheet2['C']:
                        o += 1
                        if o == Index_check:
                            Helperfunctions.Appendsheetall(WorkSheet1, r.value, 4)
                    loginin.destroy()
                    GUI.Add_Profile(self)
                else:
                    loginin.destroy()
        if Check == 0:
            messagebox.showinfo("No username found !", "Sign-up first to become a part of Facebook.")
            loginin.destroy()
        Program_execution()
    def Search(self, search, searching):
        Helperfunctions.Checking_blocking(search.get())
        if WorkSheet2.cell(row=3, column=7).value == 1:
            Helperfunctions.You_Blocked(search.get())
            if WorkSheet2.cell(row=4, column=7).value == 1:
                check = 0
                Index_check = 0
                for name in range(len(WorkSheet1['C'])):
                    Index_check += 1
                    if WorkSheet1['C'][name].value == search.get():
                        Z = WorkSheet1.cell(row=name + 1, column=3).value
                        A = WorkSheet1.cell(row=name + 1, column=4).value
                        B = WorkSheet1.cell(row=name + 1, column=6).value
                        C = WorkSheet1.cell(row=name + 1, column=7).value
                        D = WorkSheet1.cell(row=name + 1, column=8).value
                        E = WorkSheet1.cell(row=name + 1, column=9).value
                        F = WorkSheet1.cell(row=name + 1, column=10).value
                        G = WorkSheet1.cell(row=name + 1, column=11).value
                        H = WorkSheet1.cell(row=name + 1, column=13).value
                        searching.destroy()
                        GUI.See_Serching(self, Z, A, B, C, D, E, F, G, H)
                        check += 1
                if check == 0:
                    messagebox.showinfo("Not Found", "This user is not found into the facebook.")
                    searching.destroy()
            else:
                messagebox.showinfo("You blocked this user", "Unblock this user first !")
                searching.destroy()
        else:
            messagebox.showinfo("You are Blocked ", "You are blocked by this User !")
            searching.destroy()
        Program_execution()
    def Friendrequest(self, name, request):
        if name.get() != WorkSheet2.cell(row=2, column=7).value:
            Helperfunctions.Checking_blocking(name.get())
            if WorkSheet2.cell(row=3, column=7).value == 1:
                Helperfunctions.You_Blocked(name.get())
                if WorkSheet2.cell(row=4, column=7).value == 1:
                    check = 0
                    Index_check = 0
                    for i in WorkSheet3['1']:
                        Index_check += 1
                        if i.value == name.get():
                            Helperfunctions.Appendsheetall(WorkSheet3, WorkSheet2.cell(row=2, column=7).value,
                                                           Index_check)
                            messagebox.showinfo("Friend Request", "Okay! your request has been sent.")
                            for g in range(len(WorkSheet2['D'])):
                                if WorkSheet2['D'][g].value == name.get():
                                    Helperfunctions.Friendrequeststatus(True, g + 1)
                            check += 1
                    if check == 0:
                        messagebox.showinfo("No found", "This username does not exist into the Facebook.")
                else:
                    messagebox.showinfo("You Blocked this user",
                                        "Only unblocked person can recieve your friend request.")
            else:
                messagebox.showinfo("You are blocked.", "This person has blocked you.")
        else:
            messagebox.showinfo("Warning !", "You cannot send friend-request to yourself.")
        request.destroy()
        Program_execution()
    def Logout(self, logging_out):
        messagebox.showinfo("Logout Success !",WorkSheet2.cell(row=2, column=7).value+" logout.")
        WorkSheet2.cell(row=2, column=7).value = None
        Workbook1.save("Project2.xlsx")
        logging_out.destroy()
        Program_execution()
    def Createpost(self, post, posting):
        Index_check = 0
        for i in WorkSheet5['A']:
            Index_check += 1
            if i.value == WorkSheet2.cell(row=2, column=7).value:
                o = 0
                while o != Index_check:
                    o += 1
                WorkSheet5.cell(row=o, column=2).value = post.get()
                Workbook1.save("Project2.xlsx")
                messagebox.showinfo("Post Uploaded", "Okay! your post has been uploaded.")
                posting.destroy()
        Program_execution()
    def Messages(self, Name, message, messeging):
        if Name.get() != WorkSheet2.cell(row=2, column=7).value:
            Helperfunctions.Checking_blocking(Name.get())
            if WorkSheet2.cell(row=3, column=7).value == 1:
                Helperfunctions.You_Blocked(Name.get())
                if WorkSheet2.cell(row=4, column=7).value == 1:
                    member = False
                    for i in range(len(WorkSheet4['1'])):
                        if WorkSheet4['1'][i].value == Name.get():
                            k = 1
                            while k <= WorkSheet4.max_row:
                                a = WorkSheet4.cell(row=k, column=i + 1).value
                                if a == WorkSheet2.cell(row=2, column=7).value:
                                    member = True
                                k += 1
                    if member == True:
                        Helperfunctions.Appendsheetall(WorkSheet6, message.get(), 1)
                        Helperfunctions.Appendsheetall(WorkSheet6, WorkSheet2.cell(row=2, column=7).value, 2)
                        Helperfunctions.Appendsheetall(WorkSheet6, Name.get(), 3)
                        messagebox.showinfo("Message Sent", "Okay! Your message has been sent.")
                    else:
                        messagebox.showinfo("Message not sent", "This person is not your friend !")
                else:
                    messagebox.showinfo("You blocked this user", "To send message,first unblock this user !")
            else:
                messagebox.showinfo("You are blocked", "This User has blocked you !")
        else:
            messagebox.showinfo("Sending Ownself", "You cannot send message to yourself.")
        messeging.destroy()
        Program_execution()
    def Change_YourProfile(self, profiling):
        profiling.destroy()
        for i in range(len(WorkSheet1['C'])):
            if WorkSheet2.cell(row=2, column=7).value == WorkSheet1['C'][i].value:
                WorkSheet1.cell(row=i + 1, column=7).value = None
                WorkSheet1.cell(row=i + 1, column=8).value = None
                WorkSheet1.cell(row=i + 1, column=9).value = None
                WorkSheet1.cell(row=i + 1, column=10).value = None
                WorkSheet1.cell(row=i + 1, column=11).value = None
                WorkSheet1.cell(row=i + 1, column=13).value = None
                Workbook1.save("Project2.xlsx")
                GUI().Add_Profile()
    def Notify_Friend_request(self, Notify):
        for p in range(len(WorkSheet2['D'])):
            if WorkSheet2['D'][p].value == WorkSheet2.cell(row=2, column=7).value:
                Helperfunctions.Friendrequeststatus(False, p + 1)
        for i in range(len(WorkSheet3['1'])):
            if WorkSheet3['1'][i].value == WorkSheet2.cell(row=2, column=7).value:
                friendcheck = False
                k = 2
                while WorkSheet3.cell(row=k, column=i + 1).value != None:
                    username = WorkSheet3.cell(row=k, column=i + 1).value
                    friendcheck = True
                    Notify.destroy()
                    GUI().accept_reject(username)
                    k += 1
                k = 2
                while k <= WorkSheet3.max_row:
                    WorkSheet3.cell(row=k, column=i + 1).value = None
                    k += 1
                Workbook1.save('Project2.xlsx')
                if friendcheck == False:
                    messagebox.showinfo("No request found", "You have no friend request yet !")
                    Notify.destroy()
        Program_execution()
    def Notify_PostComment(self, notity):
        for p in range(len(WorkSheet2['D'])):
            if WorkSheet2['D'][p].value == WorkSheet2.cell(row=2, column=7).value:
                Helperfunctions.Commentstatus(False, p + 1)
        check = 0
        post = False
        for i in WorkSheet5['A']:
            check += 1
            if WorkSheet2.cell(row=2, column=7).value == i.value:
                comment = False
                o = 0
                for r in WorkSheet5['B']:
                    o += 1
                    if o == check and r.value != None:
                        if WorkSheet5.cell(row=o, column=3).value != None:
                            k = 3
                            while k <= WorkSheet5.max_column:
                                if WorkSheet5.cell(row=check, column=k).value != None:
                                    messagebox.showinfo("Your post: " + r.value,
                                                        WorkSheet5.cell(row=check, column=k).value)
                                k += 1
                                comment = True
                        else:
                            messagebox.showinfo("Your post: " + r.value, "There is no comment on your post.")
                            notity.destroy()
                        post = True
                        if comment == True:
                            notity.destroy()
                            GUI().commentreply(1)
        if post == False:
            messagebox.showinfo("No post", "You have no post till now.")
            notity.destroy()
        Program_execution()
    def Notify_Messages(self, notifi):
        message = False
        for i in range(len(WorkSheet6['C'])):
            if WorkSheet6['C'][i].value == WorkSheet2.cell(row=2, column=7).value:
                messagebox.showinfo(" From: " + WorkSheet6['B'][i].value, "Message: " + WorkSheet6['A'][i].value)
                message = True
                Helperfunctions.Appendsheetall(WorkSheet6, WorkSheet6['A'][i].value, 6)
                Helperfunctions.Appendsheetall(WorkSheet6, WorkSheet6['B'][i].value, 7)
                Helperfunctions.Appendsheetall(WorkSheet6, WorkSheet6['C'][i].value, 8)
                WorkSheet6['A'][i].value = None
                WorkSheet6['B'][i].value = None
                WorkSheet6['C'][i].value = None
                Workbook1.save('Project2.xlsx')
        if message == True:
            notifi.destroy()
            GUI().commentreply(2)
        else:
            messagebox.showinfo("No new Message", "You have no new message.")
            notifi.destroy()
        Program_execution()
    def Accept_Reject(self, reply, user, accept):
        if reply.get() == "1":
            check = 0
            for m in range(len(WorkSheet4['1'])):
                check += 1
                if WorkSheet4['1'][m].value == WorkSheet2.cell(row=2, column=7).value:
                    Helperfunctions.Appendsheetall(WorkSheet4, user, check)
                    Helperfunctions.Appendsheetall(WorkSheet7, user, check)
            check1 = 0
            for j in range(len(WorkSheet4['1'])):
                check1 += 1
                if WorkSheet4['1'][j].value == user:
                    Helperfunctions.Appendsheetall(WorkSheet4, WorkSheet2.cell(row=2, column=7).value, check1)
                    Helperfunctions.Appendsheetall(WorkSheet7, WorkSheet2.cell(row=2, column=7).value, check1)
            messagebox.showinfo("Request Accepted !", "Okay ! This user added in your privacy list.")
        else:
            messagebox.showinfo("Requset Rejected", "Okay! you rejected the friend request.")
        accept.destroy()
    def ReadAllmessages(self, ending):
        message = False
        for i in range(len(WorkSheet6['H'])):
            if WorkSheet6['H'][i].value == WorkSheet2.cell(row=2, column=7).value:
                messagebox.showinfo("From: " + WorkSheet6['G'][i].value, "Message: " + WorkSheet6['F'][i].value)
                message = True
        if message == True:
            ending.destroy()
            GUI().commentreply(2)
        else:
            messagebox.showinfo("No Message", "Your message folder is empty yet !")
            ending.destroy()
        Program_execution()
    def Privacy_Adding(self, entermember, adding):
        if entermember.get() == WorkSheet2.cell(row=2, column=7).value:
            messagebox.showinfo("Warning !","You can add your friends here, not yourself !")
        else:
            for i in range(len(WorkSheet4['1'])):
                if WorkSheet4['1'][i].value == WorkSheet2.cell(row=2, column=7).value:
                    entry = False
                    k = 2
                    while WorkSheet4.cell(row=k, column=i + 1).value != None:
                        a = WorkSheet4.cell(row=k, column=i + 1).value
                        if a == entermember.get():
                            Helperfunctions.Appendsheetall(WorkSheet7, entermember.get(), i + 1)
                            messagebox.showinfo("Member Added","Okay ! " + entermember.get() + " added in your privacy list.")
                            entry = True
                            break
                        k += 1
                    if entry == False:
                        messagebox.showinfo("Cannot be Added !",entermember.get()+" is not your friend.")
        adding.destroy()
        Program_execution()
    def Privacy_Removing(self, removemember, removing):
        for i in range(len(WorkSheet7['1'])):
            if WorkSheet7['1'][i].value == WorkSheet2.cell(row=2, column=7).value:
                entry = False
                k = 2
                while k <= WorkSheet7.max_row:
                    if WorkSheet7.cell(row=k, column=i + 1).value == removemember.get():
                        WorkSheet7.cell(row=k, column=i + 1).value = None
                        Workbook1.save("Project2.xlsx")
                        messagebox.showinfo("Member Removed","Okay ! " + removemember.get() + " has been removed from your privacy list.")
                        removing.destroy()
                        entry = True
                        break
                    k += 1
                if entry == False:
                    messagebox.showinfo("No found !",removemember.get()+" is not exist into your privacy list.")
                    removing.destroy()
        Program_execution()
    def User_Blocked(self, entermember, blocking):
        entry = False
        for i in range(len(WorkSheet2['D'])):
            if WorkSheet2['D'][i].value == entermember.get():
                entry = True
                break
        if entry == True:
            if entermember.get() == WorkSheet2.cell(row=2, column=7).value:
                messagebox.showinfo("Warning !","You cannot blocked yourself !")
            else:
                for b in range(len(WorkSheet8['1'])):
                    if WorkSheet8['1'][b].value == WorkSheet2.cell(row=2, column=7).value:
                        Helperfunctions.Appendsheetall(WorkSheet8, entermember.get(), b + 1)
                        messagebox.showinfo("User Blocked", entermember.get()+" has been blocked.")
        else:
            messagebox.showinfo("No Found !", "No any member found by this name.")
        blocking.destroy()
        Program_execution()
    def User_Unbloked(self, removemember, unblocking):
        entry = False
        for i in range(len(WorkSheet8['1'])):
            if WorkSheet8['1'][i].value == WorkSheet2.cell(row=2, column=7).value:
                k = 2
                while k <= WorkSheet8.max_row:
                    if WorkSheet8.cell(row=k, column=i + 1).value == removemember.get():
                        WorkSheet8.cell(row=k, column=i + 1).value = None
                        Workbook1.save("Project2.xlsx")
                        messagebox.showinfo("User Removed",removemember.get() + " has been removed from your block_members list.")
                        unblocking.destroy()
                        entry = True
                        break
                    k += 1
        if entry == False:
            messagebox.showinfo("User not exist", "This User does not exist in your block_members list.")
            unblocking.destroy()
        Program_execution()
    def Postsbyfriends(self, named, Postingend):
        Helperfunctions.Checking_blocking(named.get())
        if WorkSheet2.cell(row=3, column=7).value == 1:
            Helperfunctions.You_Blocked(named.get())
            if WorkSheet2.cell(row=4, column=7).value == 1:
                member = False
                for i in range(len(WorkSheet7['1'])):
                    if WorkSheet7['1'][i].value == named.get():
                        k = 1
                        while k <= WorkSheet7.max_row:
                            a = WorkSheet7.cell(row=k, column=i + 1).value
                            if a == WorkSheet2.cell(row=2, column=7).value:
                                member = True
                            k += 1
                if member == True:
                    check = 0
                    for i in WorkSheet5['A']:
                        check += 1
                        if named.get() == i.value:
                            o = 0
                            for r in WorkSheet5['B']:
                                o += 1
                                if o == check and r.value != None:
                                    Postingend.destroy()
                                    posted = ("Post by the " + i.value + " is: " + r.value)
                                    GUI().Feedback_post(posted, check)
                                    for u in range(len(WorkSheet2['D'])):
                                        if WorkSheet2['D'][u].value == named.get():
                                            Helperfunctions.Commentstatus(True, u + 1)
                                            break
                                    break
                            else:
                                messagebox.showinfo("No Post", i.value + " have no post at that time.")
                                Postingend.destroy()
                else:
                    messagebox.showinfo("Post is Hidden.", "User's post only visible to his/her privacy list.")
                    Postingend.destroy()
            else:
                messagebox.showinfo("You blocked " + named.get(), "First Unblock this user !")
                Postingend.destroy()
        else:
            messagebox.showinfo("You are Blocked",named.get() + " has blocked you !")
            Postingend.destroy()
        Program_execution()
    def Postbywordsearching(self, word,abcde):
        search = False
        for wording in range(len(WorkSheet5['B'])):
            if WorkSheet5['B'][wording].value != None and word.get() in WorkSheet5['B'][wording].value:
                messagebox.showinfo(word.get() + " Found !","Post " + WorkSheet5['B'][wording].value + ", Uploaded by: " + WorkSheet5['A'][wording].value)
                abcde.destroy()
                search = True
        if search == False:
            messagebox.showinfo("No Found !", "No any post finded by this word: " + word.get())
            abcde.destroy()
        Program_execution()
    def ADD_profile(self, gender, education, work, city, country, bio, adding):
        Helperfunctions.Appendsheetall(WorkSheet1, gender.get(), 7)
        Helperfunctions.Appendsheetall(WorkSheet1, education.get(), 8)
        Helperfunctions.Appendsheetall(WorkSheet1, work.get(), 9)
        Helperfunctions.Appendsheetall(WorkSheet1, city.get(), 10)
        Helperfunctions.Appendsheetall(WorkSheet1, country.get(), 11)
        Helperfunctions.Appendsheetall(WorkSheet1, bio.get(), 13)
        messagebox.showinfo("Profile Completed", "Great! Your data has been saved.")
        adding.destroy()
    def comment_yes(self, check, ending):
        ending.destroy()
        tkW = Tk()
        tkW.config(background='grey', borderwidth=5)
        tkW.geometry('300x300')
        tkW.title("Comment")
        Label(tkW, text="Write your comment !", bg="blue", bd=10, font='Helvetica 10 bold', fg='white').place(x=30,y=30)
        Label(tkW, text="Comment", background="lightblue").place(x=30,y=30)
        com = StringVar()
        Entry(tkW, textvariable=com).grid(row=2, column=1)
        Success = partial(self.Take_comment, com, check, tkW)
        Button(tkW, text="Send", command=Success).place(x=100, y=100)
        tkW.mainloop()
    def Take_comment(self, comment, check, tkW):
        a = ("Commment: " + comment.get() + " From: " + WorkSheet2.cell(row=2, column=7).value)
        Helperfunctions.Appendforfriends(WorkSheet5, a, check)
        messagebox.showinfo("Comment Done", "Thanks for comment !")
        tkW.destroy()
    def Add_Paging(self,ad_page):
        if WorkSheet2.cell(row=4,column=7).value == None:
            WorkSheet2.cell(row=4, column=7).value=0
            Workbook1.save("Project2.xlsx")
        i = WorkSheet2.cell(row=4,column=7).value
        liste= ['Sheet9','Sheet10','Sheet11','Sheet12','Sheet13']
        page_title = Workbook1[liste[i]]
        Workbook1.save(("Project2.xlsx"))
        messagebox.showinfo("Page Created",WorkSheet2.cell(row=2,column=7).value+", You created new page.")
        page_title.cell(row=1,column=1).value = "Creater:"
        page_title.cell(row=1, column=2).value = WorkSheet2.cell(row=2,column=7).value
        i += 1
        WorkSheet2.cell(row=4, column=7).value = i
        Workbook1.save(("Project2.xlsx"))
        ad_page.destroy()
        Program_execution()

def Program_execution():
    Programe_Execute = Tk()
    Programe_Execute.title("Facebook")
    if WorkSheet2.cell(row=2, column=7).value == None:
        Programe_Execute.geometry('1300x650')
        Programe_Execute.config(background='lightblue', borderwidth=9)
        Label(Programe_Execute, text="Wellcome to Facebook", bg="white", bd=17, font='Helvetica 18 bold',fg='blue').place(x=500,y=270)
        Button(Programe_Execute, text="Signup", command=partial(GUI().Signup, Programe_Execute), bg='green',fg='white', bd=13,font='Helvetica 13 bold').place(x=460, y=360)
        Button(Programe_Execute, text="Login", command=partial(GUI().Login, Programe_Execute), bg='green',fg='white', bd=13,font='Helvetica 13 bold').place(x=750, y=360)
    else:
        Programe_Execute.geometry('1500x750')
        Programe_Execute.config(background='grey', borderwidth=9)
        Label(Programe_Execute, text="Features of Facebook", bg="white", bd=20,fg='blue', font='Helvetica 35 bold').place(x=390,y=15)
        Button(Programe_Execute, text="Friend Request", command=partial(GUI().Friend_Request, Programe_Execute),bg='lime',fg='black', bd=13,font='Helvetica 13 bold').place(x=1000, y=200)
        Button(Programe_Execute, text="Log out", command=partial(Facebook().Logout, Programe_Execute),bg='red',fg='white', bd=13,font='Helvetica 13 bold').place(x=1055, y=65)
        Button(Programe_Execute, text="Profile", command=partial(GUI().Search, Programe_Execute),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=235, y=395)
        Button(Programe_Execute, text="Create Post", command=partial(GUI().Create_post, Programe_Execute),bg='lime',fg='black', bd=13,font='Helvetica 13 bold').place(x=450,y=200)
        Button(Programe_Execute, text="Send", command=partial(GUI().Message_friend, Programe_Execute, 2),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=810, y=395)
        Button(Programe_Execute, text="Edit Profile",command=partial(Facebook().Change_YourProfile, Programe_Execute),bg='lime',fg='black', bd=13,font='Helvetica 13 bold').place(x=200, y=200)
        Button(Programe_Execute, text="Notifications", command=partial(GUI().Notifications, Programe_Execute),bg='lime',fg='black', bd=13,font='Helvetica 13 bold').place(x=730,y=200)
        Button(Programe_Execute, text="Received",command=partial(Facebook().ReadAllmessages, Programe_Execute),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=1000, y=395)
        Button(Programe_Execute, text="Add Member",command=partial(GUI().Privacyadding, Programe_Execute),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=210, y=605)
        Button(Programe_Execute, text="Remove Member",command=partial(GUI().Privacyremoving, Programe_Execute),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=410, y=605)
        Button(Programe_Execute, text="Block User", command=partial(GUI().Blocking_user, Programe_Execute),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=750,y=605)
        Button(Programe_Execute, text="Unblock User", command=partial(GUI().Unblocking_user, Programe_Execute),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=1025, y=605)
        Button(Programe_Execute, text="Post", command=partial(GUI().Post_by_friends, Programe_Execute),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=350, y=395)
        Button(Programe_Execute, text="letter/word", command=partial(GUI().Word_Searching, Programe_Execute),bg='lime',fg='black', bd=8,font='Helvetica 13 bold').place(x=435, y=395)
        Label(Programe_Execute, text="Search", bg="violet", bd=20, fg='black',font='Helvetica 26 bold').place(x=300, y=300)
        Label(Programe_Execute, text="Message", bg="violet", bd=20, fg='black', font='Helvetica 26 bold').place(x=850,y=300)
        Label(Programe_Execute, text="Post Privacy", bg="violet", bd=20, fg='black', font='Helvetica 26 bold').place(x=255,y=510)
        Label(Programe_Execute, text="Blocked System", bg="violet", bd=20, fg='black', font='Helvetica 26 bold').place(x=800, y=510)
        # Button(Programe_Execute, text="Add Page", command=partial(GUI().Add_page, Programe_Execute),bg='lime',fg='black', bd=13,font='Helvetica 13 bold').place(x=1150, y=550)
    Programe_Execute.mainloop()

class GUI:
    def Signup(self, a):
        a.destroy()
        tkWindow = Tk()
        tkWindow.config(background='grey', borderwidth=5)
        tkWindow.geometry('350x300')
        tkWindow.title("Sign-up")
        Label(tkWindow, text="Create your account !", bg="blue", font='Helvetica 15 bold', fg='white').place(x=40, y=1)
        Label(tkWindow, text="Name", background="grey", fg='black', font="Helvitica 10 bold").place(x=40, y=60)
        Name = StringVar()
        Entry(tkWindow, textvariable=Name).place(x=120, y=60)

        Label(tkWindow, text="Password", background="grey", fg='black', font="Helvitica 10 bold").place(x=40, y=80)
        Password = StringVar()
        Entry(tkWindow, textvariable=Password, show='*', fg='red').place(x=120, y=80)

        Label(tkWindow, text="Mail", background="grey", fg='black', font="Helvitica 10 bold").place(x=40, y=100)
        Mail = StringVar()
        Entry(tkWindow, textvariable=Mail).place(x=120, y=100)

        Label(tkWindow, text="Username", background="grey", fg='black', font="Helvitica 10 bold").place(x=40, y=120)
        Username = StringVar()
        Entry(tkWindow, textvariable=Username).place(x=120, y=120)

        Success = partial(Facebook().Signup, Name, Password, Mail, Username, tkWindow)
        Button(tkWindow, text="Submit !", command=Success, bg="green", bd=10, font='Helvetica 10 bold',fg='white').place(x=150, y=160)
        tkWindow.mainloop()
    def Login(self, a):
        a.destroy()
        login_entry = Tk()
        login_entry.config(background='grey', borderwidth=5)
        login_entry.geometry('350x300')
        login_entry.title("Login")
        Label(login_entry, text="Enter your credentials !", bg="blue", font='Helvetica 13 bold', fg='white').place(x=30,y=20)
        Label(login_entry, text="Username", background="grey", fg='black', font="Helvitica 10 bold").place(x=50, y=60)
        username = StringVar()
        Entry(login_entry, textvariable=username).place(x=130, y=60)
        Label(login_entry, text="Password", background="grey", fg='black', font="Helvitica 10 bold").place(x=50, y=90)
        password = StringVar()
        Entry(login_entry, textvariable=password, show='*',fg='red').place(x=130, y=90)
        Button(login_entry, text="Confirm", command=partial(Facebook().Login, username, password, login_entry), bg="green", bd=10, font='Helvetica 10 bold', fg='white').place(x=150, y=130)
        login_entry.mainloop()
    def Add_Profile(self):
        adding = Tk()
        adding.config(background='grey', borderwidth=5)
        adding.geometry('400x300')
        adding.title("Add_Profile")
        Label(adding, text="Complete your Profile !", bg="blue", bd=10, font='Helvetica 10 bold', fg='white').grid(
            row=0, column=0)
        Label(adding, text="Gender: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=1, column=0)
        gender = StringVar()
        Entry(adding, textvariable=gender).grid(row=1, column=1)
        Label(adding, text="Education: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=2, column=0)
        education = StringVar()
        Entry(adding, textvariable=education).grid(row=2, column=1)
        Label(adding, text="Work: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=3, column=0)
        work = StringVar()
        Entry(adding, textvariable=work).grid(row=3, column=1)
        Label(adding, text="City: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=4, column=0)
        city = StringVar()
        Entry(adding, textvariable=city).grid(row=4, column=1)
        Label(adding, text="Country: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=5, column=0)
        country = StringVar()
        Entry(adding, textvariable=country).grid(row=5, column=1)
        Label(adding, text="Enter Bio: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=6, column=0)
        bio = StringVar()
        Entry(adding, textvariable=bio).grid(row=6, column=1)
        Success = partial(Facebook().ADD_profile, gender, education, work, city, country, bio, adding)
        Button(adding, text="Completed", command=Success, bg="green", bd=10, font='Helvetica 8 bold', fg='white').place(x=190, y=190)
        adding.mainloop()
        Program_execution()
    def Friend_Request(self, a):
        a.destroy()
        friendy = Tk()
        friendy.config(background='grey', borderwidth=5)
        friendy.geometry('400x200')
        friendy.title("Facebook Friend_Request")
        Label(friendy, text="Write Faceboook_member name to make friend !", bg="blue", bd=10, font='Helvetica 10 bold', fg='white').place(x=0,y=0)
        Label(friendy, text="Name", background="grey", fg='black', font="Helvitica 10 bold").place(x=40,y=60)
        name = StringVar()
        Entry(friendy, textvariable=name).place(x=90,y=60)
        Success = partial(Facebook().Friendrequest, name, friendy)
        Button(friendy, text="Send", command=Success, bg="green", bd=10, font='Helvetica 10 bold',fg='white').place(x=150, y=100)
        Button(friendy, text="Exit", command=partial(Helperfunctions().OFF_searching, friendy), bg='green', fg='white',bd=8, font='Helvetica 9 bold').place(x=0,y=150)
        friendy.mainloop()
    def Search(self, a):
        a.destroy()
        Searchbar = Tk()
        Searchbar.config(background='grey', borderwidth=5)
        Searchbar.geometry('350x200')
        Searchbar.title("Search Profile")
        Label(Searchbar, text="    ", background="grey").grid(row=0, column=0)
        Label(Searchbar, text="Search Profile by username", bg="blue", bd=10, font='Helvetica 10 bold', fg='white').grid(row=0, column=1)
        Label(Searchbar, text="    ", background="grey").grid(row=1, column=0)
        Label(Searchbar, text="User Name", background="grey", fg='black', font="Helvitica 10 bold").grid(row=2, column=0)
        search = StringVar()
        Entry(Searchbar, textvariable=search).grid(row=2, column=1)
        Success = partial(Facebook().Search, search, Searchbar)
        Button(Searchbar, text="Search", command=Success, bg="green", bd=10, font='Helvetica 10 bold',fg='white').place(x=100, y=100)
        Button(Searchbar, text="Exit", command=partial(Helperfunctions().OFF_searching, Searchbar), bg='green', fg='white',bd=8, font='Helvetica 9 bold').place(x=0, y=150)
        Searchbar.mainloop()
    def See_Serching(self, Z, A, B, C, D, E, F, G, H):
        SEEsearching = Tk()
        SEEsearching.config(background='grey', borderwidth=5)
        SEEsearching.geometry('400x400')
        SEEsearching.title("Person Data")
        Label(SEEsearching, text=" ", background="grey").grid(row=1, column=0)
        Label(SEEsearching, text="User Details", background="white", fg='blue', font="Helvitica 11 bold").grid(row=1, column=1)
        Label(SEEsearching, text="Username: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=2, column=0)
        Label(SEEsearching, text=Z, background="grey", fg='black', font="Helvitica 10 bold").grid(row=2, column=1)
        Label(SEEsearching, text="Gmail: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=3, column=0)
        Label(SEEsearching, text=A, background="grey", fg='black', font="Helvitica 10 bold").grid(row=3, column=1)
        Label(SEEsearching, text="Joining Date: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=4, column=0)
        Label(SEEsearching, text=B, background="grey", fg='black', font="Helvitica 10 bold").grid(row=4, column=1)
        Label(SEEsearching, text="Gender: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=5, column=0)
        Label(SEEsearching, text=C, background="grey", fg='black', font="Helvitica 10 bold").grid(row=5, column=1)
        Label(SEEsearching, text="Education: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=6, column=0)
        Label(SEEsearching, text=D, background="grey", fg='black', font="Helvitica 10 bold").grid(row=6, column=1)
        Label(SEEsearching, text="Work: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=7, column=0)
        Label(SEEsearching, text=E, background="grey", fg='black', font="Helvitica 10 bold").grid(row=7, column=1)
        Label(SEEsearching, text="City: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=9, column=0)
        Label(SEEsearching, text=F, background="grey", fg='black', font="Helvitica 10 bold").grid(row=9, column=1)
        Label(SEEsearching, text="Country: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=10, column=0)
        Label(SEEsearching, text=G, background="grey", fg='black', font="Helvitica 10 bold").grid(row=10, column=1)
        Label(SEEsearching, text="User_Bio: ", background="grey", fg='black', font="Helvitica 10 bold").grid(row=11, column=0)
        Label(SEEsearching, text=H, background="grey", fg='black', font="Helvitica 10 bold").grid(row=11, column=1)
        Button(SEEsearching, text="Exit", command=partial(Helperfunctions().OFF_searching, SEEsearching), bg='green', fg='white',bd=8, font='Helvetica 9 bold').place(x=0,y=350)
        SEEsearching.mainloop()
    def Create_post(self, a):
        a.destroy()
        posting = Tk()
        posting.config(background='grey', borderwidth=5)
        posting.geometry('400x300')
        posting.title("Create Post")
        Label(posting, text="Wellcome to Post creation.", bg="blue", bd=10, font='Helvetica 10 bold', fg='white').place(x=100,y=0)
        Label(posting, text="Write your post here:", background="grey", fg='black', font="Helvitica 10 bold").place(x=30,y=60)
        post = StringVar()
        Entry(posting, textvariable=post).place(x=175,y=60)
        Success = partial(Facebook().Createpost, post, posting)
        Button(posting, text="Create", command=Success, bg="green", bd=10, font='Helvetica 10 bold',fg='white').place(x=180, y=110)
        Button(posting, text="Exit", command=partial(Helperfunctions().OFF_searching, posting), bg='green', fg='white',bd=8, font='Helvetica 9 bold').place(x=0,y=250)
        posting.mainloop()
    def Message_friend(self, a, y):
        a.destroy()
        messeging = Tk()
        messeging.config(background='lightblue', borderwidth=5)
        messeging.geometry('500x500')
        if y == 1:
            messeging.title("Comment Reply")
            Label(messeging, text="Wellcome to Comment_Reply", background="lightblue").grid(row=0, column=0)
            Label(messeging, text="Friend Name", background="lightblue").grid(row=1, column=0)
            Name = StringVar()
            Entry(messeging, textvariable=Name).grid(row=1, column=1)
            Label(messeging, text="Reply", background="lightblue").grid(row=2, column=0)
            Message = StringVar()
            Entry(messeging, textvariable=Message).grid(row=2, column=1)
            Success = partial(Facebook().Messages, Name, Message, messeging)
            Button(messeging, text="Send", command=Success).place(x=100, y=100)
        else:
            messeging.title("Facebook Message")
            Label(messeging, text="Wellcome to Facebook_Message", background="lightblue").grid(row=0, column=0)
            Label(messeging, text="Friend Name", background="lightblue").grid(row=1, column=0)
            Name = StringVar()
            Entry(messeging, textvariable=Name).grid(row=1, column=1)
            Label(messeging, text="Message", background="lightblue").grid(row=2, column=0)
            Message = StringVar()
            Entry(messeging, textvariable=Message).grid(row=2, column=1)
            Success = partial(Facebook().Messages, Name, Message, messeging)
            Button(messeging, text="Send", command=Success).place(x=100, y=100)
        messeging.mainloop()
    def Notifications(self, a):
        a.destroy()
        Notity = Tk()
        Notity.config(background='grey', borderwidth=5)
        Notity.geometry('300x300')
        Notity.title("Notifications")
        Label(Notity, text="What you want to check ", bg="blue", bd=10, font='Helvetica 13 bold', fg='white').place(x=0,y=0)
        b = partial(Facebook().Notify_Friend_request, Notity)
        c = partial(Facebook().Notify_PostComment, Notity)
        d = partial(Facebook().Notify_Messages, Notity)
        Button(Notity, text="Friends Request", command=b,bg='lime',fg='black', bd=8,font='Helvetica 10 bold').place(x=50,y=50)
        Button(Notity, text="Post Comment", command=c,bg='lime',fg='black', bd=8,font='Helvetica 10 bold').place(x=50,y=100)
        Button(Notity, text="Message", command=d,bg='lime',fg='black', bd=8,font='Helvetica 10 bold').place(x=50,y=150)
        Button(Notity, text="Exit", command=partial(Helperfunctions().OFF_searching,Notity), bg='green', fg='white', bd=8, font='Helvetica 9 bold').place(x=0,y=250)
        Notity.mainloop()
    def accept_reject(self, person):
        accept = Tk()
        accept.config(background='grey', borderwidth=5)
        accept.geometry('400x250')
        accept.title("Friend_Request")
        Label(accept, text=person+" wants you to make frined...", bg="white", fg='blue', font="Helvitica 11 bold").place(x=60, y=20)
        Label(accept, text=" Make Friend: 1  <<---->>  No: Any other key ", bg="grey", fg='black', font="Helvitica 11 bold").place(x=20, y=70)
        reply = StringVar()
        Entry(accept, textvariable=reply).place(x=100, y=100)
        Success = partial(Facebook().Accept_Reject, reply, person, accept)
        Button(accept, text="Confirm", command=Success, bg="green", bd=10, font='Helvetica 10 bold',fg='white').place(x=140, y=145)
        accept.mainloop()
    def commentreply(self, give):
        c_reply = Tk()
        c_reply.config(bg='grey', borderwidth=5)
        c_reply.geometry('300x200')
        c_reply.title("Reply")
        Label(c_reply, text="Reply by message: ", background="white", fg='blue', font="Helvitica 11 bold").place(x=40, y=30)
        Button(c_reply, text="Yes", command=partial(GUI().Message_friend, c_reply, give), bg='green', fg='white',bd=8, font='Helvetica 9 bold').place(x=50, y=80)
        Button(c_reply, text="No", command=partial(Helperfunctions().OFF_searching, c_reply), bg='red', fg='white',bd=8, font='Helvetica 9 bold').place(x=200, y=80)
        c_reply.mainloop()
    def Privacyadding(self, a):
        a.destroy()
        privacyadd = Tk()
        privacyadd.config(background='lightblue', borderwidth=5)
        privacyadd.geometry('500x500')
        privacyadd.title("Privacy Adding Member")
        Label(privacyadd, text="Wellcome to Privacy Adding Member Post", background="lightblue").grid(row=0, column=0)
        Label(privacyadd, text="Friend Name", background="lightblue").grid(row=1, column=0)
        entrymemeber = StringVar()
        Entry(privacyadd, textvariable=entrymemeber).grid(row=1, column=1)
        Success = partial(Facebook().Privacy_Adding, entrymemeber, privacyadd)
        Button(privacyadd, text="Add_User", command=Success).place(x=100, y=100)
        privacyadd.mainloop()
    def Privacyremoving(self, a):
        a.destroy()
        privacyrem = Tk()
        privacyrem.config(background='lightblue', borderwidth=5)
        privacyrem.geometry('500x500')
        privacyrem.title("Privacy Removing Member")
        Label(privacyrem, text="Wellcome to Privacy Removing Member Post", background="lightblue").grid(row=0, column=0)
        Label(privacyrem, text="Friend Name", background="lightblue").grid(row=1, column=0)
        removingmemeber = StringVar()
        Entry(privacyrem, textvariable=removingmemeber).grid(row=1, column=1)
        Success = partial(Facebook().Privacy_Removing, removingmemeber, privacyrem)
        Button(privacyrem, text="Remove_User", command=Success).place(x=100, y=100)
        privacyrem.mainloop()
    def Blocking_user(self, a):
        a.destroy()
        blockinguser = Tk()
        blockinguser.config(background='lightblue', borderwidth=5)
        blockinguser.geometry('300x300')
        blockinguser.title("Block_Member")
        Label(blockinguser, text="Wellcome to Block_Member", background="lightblue").grid(row=0, column=0)
        Label(blockinguser, text="User Name", background="lightblue").grid(row=1, column=0)
        entrymember = StringVar()
        Entry(blockinguser, textvariable=entrymember).grid(row=1, column=1)
        Success = partial(Facebook().User_Blocked, entrymember, blockinguser)
        Button(blockinguser, text="Block_User", command=Success).place(x=100, y=100)
        blockinguser.mainloop()
    def Unblocking_user(self, a):
        a.destroy()
        unblockinguser = Tk()
        unblockinguser.config(background='lightblue', borderwidth=5)
        unblockinguser.geometry('300x300')
        unblockinguser.title("Unblock_Member")
        Label(unblockinguser, text="Wellcome to Unblock Member", background="lightblue").grid(row=0, column=0)
        Label(unblockinguser, text="Blocked_user", background="lightblue").grid(row=1, column=0)
        entrymember = StringVar()
        Entry(unblockinguser, textvariable=entrymember).grid(row=1, column=1)
        Success = partial(Facebook().User_Unbloked, entrymember, unblockinguser)
        Button(unblockinguser, text="Unblock_User", command=Success).place(x=100, y=100)
        unblockinguser.mainloop()
    def Feedback_post(self, posted, check):
        feedback = Tk()
        feedback.config(background='grey', borderwidth=5)
        feedback.geometry('600x250')
        feedback.title("Friend Post")
        Label(feedback, text=posted, background="white", fg='blue', font="Helvitica 11 bold").place(x=20, y=20)
        Label(feedback, text="Comment on the Post:", background="grey", fg='black', font="Helvitica 11 bold").place(x=220, y=120)
        bc = partial(Facebook().comment_yes, check, feedback)
        de = partial(Helperfunctions().OFF_searching, feedback)
        Button(feedback, text="Yes", command=bc, bg='green', fg='white',bd=8, font='Helvetica 9 bold').place(x=190, y=150)
        Button(feedback, text="No", command=de, bg='red', fg='white',bd=8, font='Helvetica 9 bold').place(x=370, y=150)
        feedback.mainloop()
    def Post_by_friends(self, a):
        a.destroy()
        abcde = Tk()
        abcde.config(background='grey', borderwidth=5)
        abcde.geometry('500x300')
        abcde.title("Post_Friend")
        Label(abcde, text="Here you can see your Friends_post", bg="blue", bd=10, font='Helvetica 10 bold', fg='white').place(x=100,y=10)
        Label(abcde, text="Friend Name", background="grey", fg='black', font="Helvitica 10 bold").place(x=125,y=70)
        username = StringVar()
        Entry(abcde, textvariable=username).place(x=225,y=70)
        Success = partial(Facebook().Postsbyfriends, username, abcde)
        Button(abcde, text="Check", command=Success, bg="green", bd=10, font='Helvetica 10 bold',fg='white').place(x=200, y=130)
        Button(abcde, text="Exit", command=partial(Helperfunctions().OFF_searching, abcde), bg='green',fg='white', bd=8, font='Helvetica 9 bold').place(x=0, y=250)

        abcde.mainloop()
    def Word_Searching(self, a):
        a.destroy()
        ab = Tk()
        ab.config(background='lightblue', borderwidth=5)
        ab.geometry('400x400')
        ab.title("Word Serching in posts")
        Label(ab, text="Type any word to search.", background="lightblue").grid(row=0, column=0)
        Label(ab, text="Word", background="lightblue").grid(row=1, column=0)
        us = StringVar()
        Entry(ab, textvariable=us).grid(row=1, column=1)
        Success = partial(Facebook().Postbywordsearching, us, ab)
        Button(ab, text="Check", command=Success).place(x=100, y=100)
        ab.mainloop()
    def Add_page(self,a):
        a.destroy()
        ad_page = Tk()
        ad_page.config(background='lightblue', borderwidth=5)
        ad_page.geometry('500x500')
        ad_page.title("Add Page")
        Label(ad_page, text="Wellcome to Create Page !", background="lightblue").grid(row=0,column=0)
        Label(ad_page, text="Page_Title's Name", background="lightblue").grid(row=1,column=0)
        pagetitle = StringVar()
        Entry(ad_page, textvariable=pagetitle).grid(row=1,column=1)
        Success = partial(Facebook().Add_Paging,ad_page)
        Button(ad_page, text="Create", command=Success).place(x=100, y=100)
        ad_page.mainloop()

Program_execution()
